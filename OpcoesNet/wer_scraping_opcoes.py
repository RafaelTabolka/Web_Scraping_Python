import pandas as pd
import requests
from openpyxl.styles import NamedStyle
from openpyxl import load_workbook
import datetime

def ListarOpcoes(ticker, vencimento):    
    url = f'https://opcoes.net.br/listaopcoes/completa?cache={vencimento}_16h1&au=False&uinhc=0&idLista=ML&idAcao={ticker}&listarVencimentos=true&cotacoes=true'
    r = requests.get(url).json()
    l = [[ticker, vencimento, i[0].split('_')[0], i[2], i[3], i[5], i[8]] for i in r['data']['cotacoesOpcoes']]
    return pd.DataFrame(l, columns= ['ticker', 'vencimento', 'ativo', 'tipo', 'modelo', 'strike', 'preco'])

def ListarTudo(ativoObj):
    url = f'https://opcoes.net.br/listaopcoes/completa?cache=2023-5-22_16h1&au=False&uinhc=0&idLista=ML&idAcao={ativoObj}&listarVencimentos=true&cotacoes=true'
    r = requests.get(url).json()
    vencimentos = [i ['value'] for i in r ['data']['vencimentos']]
    df = pd.concat([ListarOpcoes(ativoObj, vencimento) for vencimento in vencimentos])
    return df

dados = ListarTudo("PETR4")
allDados = pd.DataFrame(dados)

allDados.to_excel("Arquivo.xlsx",index = False)

tabela = load_workbook("Arquivo.xlsx")
tabelaAtiva = tabela.active

dateStyle = NamedStyle(name='vencimento',number_format="dd/mm/yyyy")

for i in range(2, tabelaAtiva.max_row + 1):
    target = f'B{i}'

    value = datetime.datetime.strptime(
        f'{tabelaAtiva[target].value}', r'%Y-%m-%d')
    tabelaAtiva[target].value = value
    tabelaAtiva[target].style = dateStyle

tabela.save("Arquivo.xlsx")

tabela.close()
